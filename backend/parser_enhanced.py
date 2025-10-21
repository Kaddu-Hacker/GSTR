"""Enhanced parser with canonical normalization and auto-mapping"""

import openpyxl
import csv
import io
import zipfile
from typing import List, Dict, Tuple, Optional
from decimal import Decimal
from datetime import datetime
import logging

from models_canonical import (
    CanonicalInvoiceLine, FileType, DocumentType, GSTRSection,
    HeaderMapping, FileInfo
)
from decimal_utils import parse_money, compute_tax
from auto_mapper import HeaderMatcher
from canonical_fields import CANONICAL_FIELDS, DOCUMENT_TYPES
from utils import normalize_state_to_code

logger = logging.getLogger(__name__)


class EnhancedFileParser:
    """Enhanced parser with auto-mapping and canonical normalization"""
    
    def __init__(self, seller_state_code: str = "27"):
        self.seller_state_code = seller_state_code
        self.header_matcher = HeaderMatcher()
    
    def extract_files_from_zip(self, zip_content: bytes) -> List[Tuple[str, bytes]]:
        """Extract files from ZIP archive"""
        files = []
        try:
            with zipfile.ZipFile(io.BytesIO(zip_content)) as zf:
                for filename in zf.namelist():
                    if filename.endswith(('.xlsx', '.xls', '.csv')) and not filename.startswith('__MACOSX'):
                        content = zf.read(filename)
                        files.append((filename, content))
        except Exception as e:
            logger.error(f"Error extracting ZIP: {str(e)}")
        return files
    
    def read_headers(self, content: bytes, filename: str) -> Optional[List[str]]:
        """Read first row headers from file"""
        try:
            if filename.endswith('.csv'):
                # CSV file
                text = content.decode('utf-8-sig')
                reader = csv.reader(io.StringIO(text))
                headers = next(reader)
                return [h.strip() for h in headers]
            else:
                # Excel file
                wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
                ws = wb.active
                headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                wb.close()
                return [str(h).strip() if h else "" for h in headers]
        except Exception as e:
            logger.error(f"Error reading headers from {filename}: {str(e)}")
            return None
    
    def detect_and_classify_file(
        self,
        filename: str,
        content: bytes
    ) -> FileInfo:
        """
        Detect file type and suggest mappings
        
        Returns:
            FileInfo with mapping suggestions
        """
        headers = self.read_headers(content, filename)
        
        if not headers:
            return FileInfo(
                filename=filename,
                file_type=FileType.UNKNOWN,
                detected=False,
                needs_mapping=True
            )
        
        # Try auto-mapping
        mappings = self.header_matcher.map_headers(headers)
        suggested_section = self.header_matcher.suggest_section(headers)
        
        # Calculate mapping confidence
        if mappings:
            avg_confidence = sum(m.confidence for m in mappings.values()) / len(mappings)
        else:
            avg_confidence = 0.0
        
        # Detect file type based on filename and mappings
        file_type = self._detect_file_type(filename, headers, suggested_section)
        
        # Count rows
        row_count = self._count_rows(content, filename)
        
        # Determine if mapping is needed
        needs_mapping = avg_confidence < 0.75
        
        return FileInfo(
            filename=filename,
            file_type=file_type,
            detected=avg_confidence >= 0.75,
            row_count=row_count,
            columns=headers,
            mapping_confidence=avg_confidence,
            needs_mapping=needs_mapping
        )
    
    def _detect_file_type(self, filename: str, headers: List[str], suggested_section: Optional[str]) -> FileType:
        """Detect file type from filename and headers"""
        filename_lower = filename.lower()
        
        # Check filename patterns
        if "tcs_sales_return" in filename_lower or "sales_return" in filename_lower:
            return FileType.TCS_SALES_RETURN
        elif "tcs_sales" in filename_lower:
            return FileType.TCS_SALES
        elif "tax_invoice" in filename_lower or "invoice_details" in filename_lower:
            return FileType.TAX_INVOICE
        elif "credit_note" in filename_lower or "credit" in filename_lower:
            return FileType.CREDIT_NOTE
        elif "debit_note" in filename_lower or "debit" in filename_lower:
            return FileType.DEBIT_NOTE
        elif "hsn" in filename_lower:
            return FileType.HSN_SUMMARY
        elif "b2b" in filename_lower:
            return FileType.B2B_INVOICES
        
        # Check by suggested section
        if suggested_section:
            if suggested_section == "b2b":
                return FileType.B2B_INVOICES
            elif suggested_section == "hsn":
                return FileType.HSN_SUMMARY
            elif suggested_section in ["b2cs", "b2cl"]:
                return FileType.TCS_SALES
        
        return FileType.UNKNOWN
    
    def _count_rows(self, content: bytes, filename: str) -> int:
        """Count rows in file"""
        try:
            if filename.endswith('.csv'):
                text = content.decode('utf-8-sig')
                return len(text.splitlines()) - 1  # Exclude header
            else:
                wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
                ws = wb.active
                count = ws.max_row - 1  # Exclude header
                wb.close()
                return count
        except:
            return 0
    
    def parse_file_with_mapping(
        self,
        content: bytes,
        filename: str,
        upload_id: str,
        mappings: Dict[str, HeaderMapping],
        file_type: FileType = FileType.UNKNOWN
    ) -> List[CanonicalInvoiceLine]:
        """
        Parse file using provided header mappings
        
        Args:
            content: File content bytes
            filename: Filename
            upload_id: Upload ID
            mappings: Dict of file_header -> HeaderMapping
            file_type: Detected file type
        
        Returns:
            List of CanonicalInvoiceLine objects
        """
        lines = []
        
        try:
            # Read all rows
            rows = self._read_all_rows(content, filename)
            
            if not rows:
                return lines
            
            headers = rows[0]
            data_rows = rows[1:]
            
            # Create reverse mapping: canonical_field -> file_header
            canonical_to_file = {m.canonical_field: file_header for file_header, m in mappings.items()}
            
            for row_idx, row_data in enumerate(data_rows):
                try:
                    # Build row dict with canonical field names
                    row = {}
                    for i, header in enumerate(headers):
                        if i < len(row_data):
                            row[header] = row_data[i]
                    
                    # Map to canonical fields
                    canonical_row = {}
                    for file_header, value in row.items():
                        if file_header in mappings:
                            canonical_field = mappings[file_header].canonical_field
                            canonical_row[canonical_field] = value
                        else:
                            canonical_row[file_header] = value  # Keep unmapped fields
                    
                    # Parse into CanonicalInvoiceLine
                    invoice_line = self._parse_row_to_canonical(
                        canonical_row,
                        upload_id,
                        file_type,
                        row
                    )
                    
                    if invoice_line:
                        lines.append(invoice_line)
                
                except Exception as e:
                    logger.error(f"Error parsing row {row_idx} in {filename}: {str(e)}")
                    continue
        
        except Exception as e:
            logger.error(f"Error parsing file {filename}: {str(e)}")
        
        return lines
    
    def _read_all_rows(self, content: bytes, filename: str) -> List[List]:
        """Read all rows from file"""
        rows = []
        
        try:
            if filename.endswith('.csv'):
                text = content.decode('utf-8-sig')
                reader = csv.reader(io.StringIO(text))
                rows = list(reader)
            else:
                wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
                ws = wb.active
                for row in ws.iter_rows(values_only=True):
                    rows.append(list(row))
                wb.close()
        except Exception as e:
            logger.error(f"Error reading rows from {filename}: {str(e)}")
        
        return rows
    
    def _parse_row_to_canonical(
        self,
        row: Dict,
        upload_id: str,
        file_type: FileType,
        raw_data: Dict
    ) -> Optional[CanonicalInvoiceLine]:
        """
        Parse row dict to CanonicalInvoiceLine with normalization
        """
        try:
            # Extract and normalize invoice number
            invoice_no_raw = str(row.get("invoice_no", "")).strip()
            if not invoice_no_raw:
                return None
            
            invoice_no_norm = invoice_no_raw.upper().strip()
            
            # Detect document type
            doc_type = self._detect_document_type(row, file_type)
            
            # Parse monetary values using Decimal
            taxable_value_raw_decimal = parse_money(row.get("taxable_value", 0))
            gst_rate_decimal = parse_money(row.get("gst_rate", 0))
            
            # Determine if return (negative value)
            is_return = file_type == FileType.TCS_SALES_RETURN or taxable_value_raw_decimal < 0
            if is_return and taxable_value_raw_decimal > 0:
                taxable_value_raw_decimal = -taxable_value_raw_decimal
            
            # Normalize state
            place_of_supply_state = row.get("place_of_supply", "")
            place_of_supply_code = normalize_state_to_code(place_of_supply_state) or self.seller_state_code
            
            # Compute tax with Decimal precision
            tax_result = compute_tax(
                taxable_value_raw_decimal,
                gst_rate_decimal,
                self.seller_state_code,
                place_of_supply_code
            )
            
            # Parse date
            invoice_date = self._parse_date(row.get("invoice_date"))
            
            # Detect GSTR section
            gstr_section = self._detect_gstr_section(row)
            
            # Create canonical line
            canonical_line = CanonicalInvoiceLine(
                upload_id=upload_id,
                invoice_no_raw=invoice_no_raw,
                invoice_no_norm=invoice_no_norm,
                doc_type=doc_type,
                invoice_date=invoice_date,
                gstin_uin=row.get("gstin_uin"),
                place_of_supply_state=place_of_supply_state,
                place_of_supply_code=place_of_supply_code,
                taxable_value_raw=str(taxable_value_raw_decimal),
                taxable_value=float(taxable_value_raw_decimal),
                gst_rate=float(gst_rate_decimal),
                computed_tax=tax_result,
                is_return=is_return,
                is_intra_state=tax_result["is_intra_state"],
                origin="meesho" if "meesho" in file_type.value.lower() else "manual",
                gstr_section=gstr_section,
                file_type=file_type,
                hsn_code=row.get("hsn_code"),
                raw_data=raw_data
            )
            
            return canonical_line
        
        except Exception as e:
            logger.error(f"Error parsing row to canonical: {str(e)}")
            return None
    
    def _detect_document_type(self, row: Dict, file_type: FileType) -> DocumentType:
        """Detect document type from row data"""
        # Check invoice_type field
        inv_type = str(row.get("invoice_type", "")).lower().strip()
        
        for key, doc_type in DOCUMENT_TYPES.items():
            if key in inv_type:
                return DocumentType(doc_type)
        
        # Check file type
        if file_type == FileType.CREDIT_NOTE:
            return DocumentType.CREDIT_NOTE
        elif file_type == FileType.DEBIT_NOTE:
            return DocumentType.DEBIT_NOTE
        
        return DocumentType.TAX_INVOICE
    
    def _detect_gstr_section(self, row: Dict) -> Optional[GSTRSection]:
        """Auto-detect GSTR-1 section for this row"""
        gstin = row.get("gstin_uin", "")
        taxable_value = parse_money(row.get("taxable_value", 0))
        
        # B2B: Has valid GSTIN
        if gstin and len(str(gstin).strip()) == 15:
            return GSTRSection.B2B
        
        # B2CL: No GSTIN and value > 2.5L
        if not gstin and taxable_value > Decimal("250000"):
            return GSTRSection.B2CL
        
        # B2CS: No GSTIN and value <= 2.5L
        if not gstin:
            return GSTRSection.B2CS
        
        return None
    
    def _parse_date(self, date_value) -> Optional[str]:
        """Parse date to ISO format YYYY-MM-DD"""
        if not date_value:
            return None
        
        try:
            if isinstance(date_value, datetime):
                return date_value.strftime("%Y-%m-%d")
            
            # Try parsing string
            date_str = str(date_value).strip()
            # Add more date format parsing as needed
            return date_str
        except:
            return None
